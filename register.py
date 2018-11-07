from openpyxl import load_workbook
import os

#setting filenameList
filenameList = 'sensor_register_file.xlsx'
#open excel 
wb = load_workbook(filenameList)
print("open filenameList:",filenameList)

#select sheet_0 ->Module
sheets = wb.sheetnames
sheet0 = sheets[0]
ws = wb[sheet0]
print("use sheet_0(Module):",sheet0)

#get sheet_0 row and column
sheet0_row = ws.max_row
sheet0_column = ws.max_column
print("sheet0_row:",sheet0_row)
print("sheet0_column:",sheet0_column)

#parsing baseAddressList
baseAddressList = list()

for row in range(3,sheet0_row+1):
    baseAddressList.append(ws.cell(row=row,column=1).value)
print("baseAddressList[0]:",baseAddressList[0])

#select sheet_1 ->Register
sheet1 = sheets[1]
wd = wb[sheet1]
print("use sheet_1(Register):",sheet1)

#get sheet_1 row and column
sheet1_row = wd.max_row
sheet1_column = wd.max_column
print("sheet1_row:",sheet1_row)
print("sheet1_column:",sheet1_column)

#get offsetList column(range start 3 to sheet1_row+1 )
offsetList = list()
for row in range(3,sheet1_row+1): 
    offsetList.append(wd.cell(row=row,column=1).value)
print(offsetList)

#get nameListList column(range start 3 to sheet1_row+1 )
nameList = list()
for row in range(3,sheet1_row+1):
    nameList.append(wd.cell(row=row,column=2).value)
print(nameList)

#get msbList column(range start 3 to sheet1_row+1 )
msbList = list()
for row in range(3,sheet1_row+1):
    msbList.append(wd.cell(row=row,column=3).value)
print(msbList)

#get lsbList column(range start 3 to sheet1_row+1 )
lsbList = list()
for row in range(3,sheet1_row+1):
    lsbList.append(wd.cell(row=row,column=4).value)
print(lsbList)

#get newmsbList column((msbList[i]-lsbList[i])+1 )
newmsbList = list()
for i in range(len(msbList)):
    newmsbList.append((msbList[i]-lsbList[i])+1)
print(newmsbList)

#get accessList column(range start 3 to sheet1_row+1 )
accessList = list()
for row in range(3,sheet1_row+1):
    accessList.append(wd.cell(row=row,column=6).value)
print(accessList)

#get defaultValueList column(range start 3 to sheet1_row+1 )
defaultValueList = list()
for row in range(3,sheet1_row+1):
    defaultValueList.append(wd.cell(row=row,column=7).value)
print(defaultValueList)

print(baseAddressList[0])

outputFilenameList = filenameList.split(".")[0] +"_parse.csv"
print(outputFilenameList)

#write output file
f = open(outputFilenameList,'w+')
for i in range(len(offsetList)):
    row_data = str(baseAddressList[0])+str(offsetList[i]).split("x")[1]+","+str(nameList[i])+","+str(newmsbList[i])+","+str(accessList[i])+","+str(defaultValueList[i])
    print(row_data)
    f.write(row_data+'\n')